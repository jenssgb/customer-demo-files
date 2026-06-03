"""Generate SAP vs TM1 month-end close demo datasets.

Creates two primary Excel datasets plus CSV exports for Analyst uploads:
- SAP_Actuals_May2026.xlsx / .csv
- TM1_Budget_Forecast_May2026.xlsx / .csv

The generated data intentionally includes finance demo issues:
- SAP vs TM1 account and cost-center mapping differences
- Account-type-aware variance logic
- Late postings, high-value manual journals, orphan TM1 planning lines
- Property/business context for Real Estate flavored close review
"""

from __future__ import annotations

import random
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

SEED = 260603
random.seed(SEED)
np.random.seed(SEED)

ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)

COMPANY_CODES = [
    ("1000", "MUC Holding GmbH", "Germany", "EUR", "DACH"),
    ("1100", "Berlin Services GmbH", "Germany", "EUR", "DACH"),
    ("1200", "Hamburg Asset Mgmt GmbH", "Germany", "EUR", "DACH"),
    ("2000", "Paris Operations SAS", "France", "EUR", "Western Europe"),
    ("2100", "Amsterdam Services BV", "Netherlands", "EUR", "Western Europe"),
    ("3000", "London Property Ltd", "United Kingdom", "GBP", "UK"),
    ("4000", "New York Ops Inc", "United States", "USD", "Americas"),
    ("5000", "Singapore Hub Pte Ltd", "Singapore", "SGD", "APAC"),
]

ACCOUNT_DEFS = [
    ("400100", "Rental income", "Revenue", "Income Statement", "Revenue"),
    ("400200", "Service charge income", "Revenue", "Income Statement", "Revenue"),
    ("401000", "Parking income", "Revenue", "Income Statement", "Revenue"),
    ("410000", "Other operating income", "Revenue", "Income Statement", "Revenue"),
    ("500100", "Property management fees", "COGS", "Income Statement", "Expense"),
    ("510000", "Repairs and maintenance", "OpEx", "Income Statement", "Expense"),
    ("511000", "Utilities electricity", "OpEx", "Income Statement", "Expense"),
    ("511100", "Utilities heating", "OpEx", "Income Statement", "Expense"),
    ("512000", "Security services", "OpEx", "Income Statement", "Expense"),
    ("513000", "Cleaning services", "OpEx", "Income Statement", "Expense"),
    ("520000", "Insurance", "OpEx", "Income Statement", "Expense"),
    ("530000", "Property tax", "OpEx", "Income Statement", "Expense"),
    ("540000", "Professional services", "OpEx", "Income Statement", "Expense"),
    ("550000", "IT software and licenses", "OpEx", "Income Statement", "Expense"),
    ("560000", "Travel and meetings", "OpEx", "Income Statement", "Expense"),
    ("570000", "Marketing leasing campaigns", "OpEx", "Income Statement", "Expense"),
    ("580000", "Bad debt expense", "OpEx", "Income Statement", "Expense"),
    ("590000", "Other operating expense", "OpEx", "Income Statement", "Expense"),
    ("600000", "Depreciation buildings", "Depreciation", "Income Statement", "Expense"),
    ("610000", "Amortization lease incentives", "Depreciation", "Income Statement", "Expense"),
    ("700000", "Interest expense", "Finance", "Income Statement", "Expense"),
    ("710000", "FX gain loss", "Finance", "Income Statement", "Mixed"),
    ("800000", "Income tax expense", "Tax", "Income Statement", "Expense"),
    ("150000", "Capex projects WIP", "Capex", "Balance Sheet", "Asset"),
    ("151000", "Tenant improvements", "Capex", "Balance Sheet", "Asset"),
]

DEPARTMENTS = [
    "Executive",
    "Finance",
    "Facilities",
    "IT",
    "HR",
    "Sales",
    "Operations",
    "Legal",
    "Marketing",
    "Asset Management",
    "Property Management",
]
REGIONS = ["DACH", "Western Europe", "UK", "Americas", "APAC"]
VENDORS = [
    "EnergiaPlus",
    "UrbanClean GmbH",
    "SecureWorks Facilities",
    "LiftPro Services",
    "LegalTrust LLP",
    "CloudSuite SE",
    "NorthBuild Contractors",
    "GreenHeat Utilities",
    "WaterWorks City",
    "RentAdjust Consulting",
    "TaxAdvisors Global",
    "LeaseOps Services",
]


def build_accounts() -> pd.DataFrame:
    accounts = pd.DataFrame(
        ACCOUNT_DEFS,
        columns=["SAP_GL_Account", "SAP_Account_Name", "Account_Group", "Statement", "Account_Type"],
    )
    accounts["TM1_Account"] = accounts["SAP_GL_Account"].map(lambda value: "A" + value[:3] + "_" + value[3:])
    accounts.loc[accounts["SAP_GL_Account"].eq("710000"), "TM1_Account"] = "A710_FX_GL"
    accounts["Variance_Direction"] = accounts["Account_Type"].map(
        {
            "Revenue": "Higher is favorable",
            "Expense": "Lower is favorable",
            "Asset": "Within approved capex envelope",
            "Mixed": "Context dependent",
        }
    )
    return accounts


def build_cost_centers() -> pd.DataFrame:
    rows = []
    for dept_index, department in enumerate(DEPARTMENTS, 1):
        for region in REGIONS:
            region_code = region.replace(" ", "")[:3].upper()
            rows.append(
                {
                    "SAP_Cost_Center": f"CC-{dept_index:02d}-{region_code}",
                    "SAP_Cost_Center_Name": f"{department} {region}",
                    "Department": department,
                    "Region": region,
                    "TM1_Cost_Center": f"{department[:3].upper()}_{region_code}",
                    "Owner": random.choice(
                        [
                            "A. Becker",
                            "S. Martin",
                            "L. Chen",
                            "M. Rossi",
                            "N. Patel",
                            "J. Schneider",
                            "E. Dubois",
                            "T. Nguyen",
                        ]
                    ),
                }
            )
    return pd.DataFrame(rows)


def build_properties() -> pd.DataFrame:
    rows = []
    asset_types = ["Office", "Retail", "Logistics", "Residential", "Mixed Use"]
    for index in range(1, 73):
        company_code, _, country, currency, region = random.choice(COMPANY_CODES)
        rows.append(
            {
                "Property_ID": f"P-{index:03d}",
                "Property_Name": random.choice(
                    ["Riverside", "Harbor", "Central", "Northgate", "Skyline", "Westpark", "Eastpoint", "Campus", "Foundry", "Summit"]
                )
                + f" {index:02d}",
                "Company_Code": company_code,
                "Country": country,
                "Region": region,
                "Currency": currency,
                "Asset_Type": random.choice(asset_types),
                "GLA_sqm": random.randint(3800, 64000),
                "Occupancy_Pct": round(random.uniform(0.78, 0.99), 3),
                "Energy_Rating": random.choice(["A", "B", "C", "D", "E"]),
                "Property_Manager": random.choice(["CBRE", "JLL", "Cushman", "Internal PM", "BNP RE", "Savills"]),
            }
        )
    return pd.DataFrame(rows)


def amount_for(account: pd.Series, property_row: pd.Series, month: pd.Timestamp) -> float:
    if account["Account_Group"] == "Revenue":
        rent_rate = random.uniform(6, 14) if property_row["Asset_Type"] == "Logistics" else random.uniform(16, 42)
        base = property_row["GLA_sqm"] * property_row["Occupancy_Pct"] * rent_rate
    elif account["Account_Group"] in ["OpEx", "COGS"]:
        base = -(property_row["GLA_sqm"] * random.uniform(0.45, 5.5))
    elif account["Account_Group"] == "Depreciation":
        base = -(property_row["GLA_sqm"] * random.uniform(0.55, 1.8))
    elif account["Account_Group"] == "Finance":
        base = -(property_row["GLA_sqm"] * random.uniform(0.2, 1.1))
    elif account["Account_Group"] == "Tax":
        base = -(property_row["GLA_sqm"] * random.uniform(0.08, 0.7))
    elif account["Account_Group"] == "Capex":
        base = property_row["GLA_sqm"] * random.uniform(0.15, 3.2)
    else:
        base = 0
    seasonal = 1.0
    if account["SAP_GL_Account"] in ["511000", "511100"] and month.month in [1, 2, 3]:
        seasonal += 0.18
    if account["SAP_GL_Account"] == "570000" and month.month in [3, 4, 5]:
        seasonal += 0.08
    return round(base * seasonal * random.uniform(0.96, 1.04), 2)


def build_tm1(accounts: pd.DataFrame, cost_centers: pd.DataFrame, properties: pd.DataFrame) -> pd.DataFrame:
    rows = []
    months = pd.date_range("2026-01-01", "2026-05-01", freq="MS")
    for month in months:
        for _, property_row in properties.iterrows():
            for _, account in accounts.iterrows():
                if account["Statement"] != "Income Statement" and random.random() > 0.25:
                    continue
                cost_center = cost_centers.sample(1, random_state=random.randint(1, 100000)).iloc[0]
                budget = amount_for(account, property_row, month)
                forecast = round(budget * random.uniform(0.94, 1.08), 2)
                rows.append(
                    {
                        "TM1_Version": "Forecast V2 May 2026",
                        "Scenario": random.choice(["Budget", "Forecast", "Rolling Forecast"]),
                        "Month": month.strftime("%Y-%m"),
                        "Entity": property_row["Company_Code"],
                        "Property_ID": property_row["Property_ID"],
                        "TM1_Cost_Center": cost_center["TM1_Cost_Center"],
                        "TM1_Account": account["TM1_Account"],
                        "SAP_GL_Account_Mapped": account["SAP_GL_Account"],
                        "Account_Group": account["Account_Group"],
                        "Account_Type": account["Account_Type"],
                        "Budget_Amount": budget,
                        "Forecast_Amount": forecast,
                        "Currency": property_row["Currency"],
                        "TM1_Last_Update": (month + pd.Timedelta(days=random.randint(15, 25))).strftime("%Y-%m-%d"),
                        "Planning_Owner": random.choice(["FP&A Germany", "FP&A Europe", "Group Controlling", "Asset Finance", "Treasury"]),
                        "Commentary_Status": random.choice(["Commented", "Missing owner comment", "Draft", "Reviewed"]),
                    }
                )
    tm1 = pd.DataFrame(rows)
    orphan_rows = []
    for index in range(140):
        source = tm1.sample(1, random_state=3000 + index).iloc[0].copy()
        source["TM1_Account"] = random.choice(["A999_OLDRENT", "A998_LOCALTAX", "A997_MISC_CAPEX"])
        source["SAP_GL_Account_Mapped"] = ""
        source["Budget_Amount"] = round(source["Budget_Amount"] * random.uniform(0.7, 1.4), 2)
        source["Forecast_Amount"] = round(source["Forecast_Amount"] * random.uniform(0.7, 1.4), 2)
        source["Commentary_Status"] = "Mapping missing"
        orphan_rows.append(source)
    return pd.concat([tm1, pd.DataFrame(orphan_rows)], ignore_index=True)


def build_sap(accounts: pd.DataFrame, cost_centers: pd.DataFrame, properties: pd.DataFrame, tm1: pd.DataFrame) -> pd.DataFrame:
    rows = []
    posting_id = 1
    sampled_plan = tm1[tm1["SAP_GL_Account_Mapped"].ne("")].sample(n=7600, random_state=42)
    for _, plan in sampled_plan.iterrows():
        account = accounts[accounts["SAP_GL_Account"].eq(plan["SAP_GL_Account_Mapped"])].iloc[0]
        cost_center_match = cost_centers[cost_centers["TM1_Cost_Center"].eq(plan["TM1_Cost_Center"])]
        sap_cost_center = cost_center_match.iloc[0]["SAP_Cost_Center"] if len(cost_center_match) else random.choice(cost_centers["SAP_Cost_Center"].tolist())
        month_start = datetime.strptime(plan["Month"] + "-01", "%Y-%m-%d")
        line_count = random.randint(1, 5)
        for line_item in range(1, line_count + 1):
            drift = random.normalvariate(1.0, 0.11)
            if account["SAP_GL_Account"] in ["510000", "511000", "540000", "570000"] and random.random() < 0.04:
                drift *= random.uniform(1.45, 2.4)
            if account["Account_Group"] == "Revenue" and random.random() < 0.03:
                drift *= random.uniform(0.62, 0.82)
            amount = plan["Forecast_Amount"] / line_count * drift
            signed_amount = amount if account["Account_Type"] == "Revenue" else -abs(amount)
            debit_credit = "H" if account["Account_Type"] == "Revenue" else "S"
            if account["Account_Type"] == "Asset":
                signed_amount = abs(amount)
                debit_credit = "S"
            posting_date = month_start + timedelta(days=random.randint(0, 34))
            journal_source = random.choices(
                ["SAP FI", "SAP CO Allocation", "SAP RE-FX", "Manual Journal", "Accrual Engine", "AP Invoice", "Depreciation Run", "Payroll Upload"],
                weights=[25, 10, 16, 7, 13, 20, 5, 4],
            )[0]
            close_risk = ""
            if posting_date.month != month_start.month:
                close_risk = "Late posting"
            if journal_source == "Manual Journal" and abs(signed_amount) > 50000:
                close_risk = "Manual high value"
            rows.append(
                {
                    "Document_Number": f"51{posting_id:08d}",
                    "Line_Item": line_item,
                    "Company_Code": plan["Entity"],
                    "Fiscal_Year": 2026,
                    "Posting_Month": plan["Month"],
                    "Posting_Date": posting_date.strftime("%Y-%m-%d"),
                    "Document_Date": (posting_date - timedelta(days=random.randint(0, 10))).strftime("%Y-%m-%d"),
                    "SAP_GL_Account": account["SAP_GL_Account"],
                    "SAP_Account_Name": account["SAP_Account_Name"],
                    "SAP_Cost_Center": sap_cost_center,
                    "Property_ID": plan["Property_ID"],
                    "Vendor": random.choice(VENDORS) if account["Account_Group"] != "Revenue" else "",
                    "Customer_or_Tenant": random.choice(["Tenant A", "Tenant B", "Tenant C", "Anchor Tenant", "Retail Tenant", "Logistics Tenant"])
                    if account["Account_Group"] == "Revenue"
                    else "",
                    "Document_Type": random.choice(["SA", "KR", "RE", "AB", "WE", "AA", "FX"]),
                    "Journal_Source": journal_source,
                    "Debit_Credit": debit_credit,
                    "Amount_Local": round(abs(signed_amount), 2),
                    "Signed_Amount": round(signed_amount, 2),
                    "Currency": plan["Currency"],
                    "Text": random.choice(
                        [
                            "Monthly posting",
                            "Accrual reversal",
                            "Lease adjustment",
                            "Utility true-up",
                            "Manual reclass",
                            "Vendor invoice",
                            "Quarterly allocation",
                            "Late invoice capture",
                        ]
                    ),
                    "Close_Risk_Flag": close_risk,
                    "Created_By": random.choice(["SAP_BATCH", "JSMITH", "MBECKER", "LCHEN", "AP_BOT", "CONTROLLING"]),
                }
            )
            posting_id += 1
    sap = pd.DataFrame(rows)
    duplicates = []
    for index, row in sap.head(60).iterrows():
        duplicate = row.copy()
        duplicate["Document_Number"] = f"59{90000000 + index}"
        duplicate["Journal_Source"] = "Manual Journal"
        duplicate["Text"] = "Manual top-side adjustment - investigate"
        duplicate["Signed_Amount"] = round(duplicate["Signed_Amount"] * random.uniform(1.8, 3.5), 2)
        duplicate["Amount_Local"] = abs(duplicate["Signed_Amount"])
        duplicate["Close_Risk_Flag"] = "Manual high value"
        duplicate["Created_By"] = "CONTROLLING"
        duplicates.append(duplicate)
    return pd.concat([sap, pd.DataFrame(duplicates)], ignore_index=True)


def build_variance(accounts: pd.DataFrame, properties: pd.DataFrame, sap: pd.DataFrame, tm1: pd.DataFrame) -> pd.DataFrame:
    sap_agg = (
        sap.groupby(["Company_Code", "Posting_Month", "Property_ID", "SAP_GL_Account"], as_index=False)["Signed_Amount"]
        .sum()
        .rename(columns={"Signed_Amount": "SAP_Actual_Amount"})
    )
    tm1_agg = (
        tm1[tm1["SAP_GL_Account_Mapped"].ne("")]
        .groupby(["Entity", "Month", "Property_ID", "SAP_GL_Account_Mapped", "Account_Group", "Account_Type"], as_index=False)
        .agg({"Budget_Amount": "sum", "Forecast_Amount": "sum"})
        .rename(columns={"Entity": "Company_Code", "Month": "Posting_Month", "SAP_GL_Account_Mapped": "SAP_GL_Account"})
    )
    variance = pd.merge(tm1_agg, sap_agg, how="outer", on=["Company_Code", "Posting_Month", "Property_ID", "SAP_GL_Account"])
    for column in ["SAP_Actual_Amount", "Budget_Amount", "Forecast_Amount"]:
        variance[column] = variance[column].fillna(0)
    variance = variance.merge(accounts[["SAP_GL_Account", "SAP_Account_Name", "Account_Group", "Account_Type"]], on="SAP_GL_Account", how="left", suffixes=("", "_map"))
    variance["Account_Group"] = variance["Account_Group"].fillna(variance["Account_Group_map"])
    variance["Account_Type"] = variance["Account_Type"].fillna(variance["Account_Type_map"])
    variance.drop(columns=["Account_Group_map", "Account_Type_map"], inplace=True)
    variance["Absolute_Variance_vs_Forecast"] = variance["SAP_Actual_Amount"] - variance["Forecast_Amount"]
    variance["Variance_Pct_vs_Forecast"] = np.where(
        variance["Forecast_Amount"].abs() > 1,
        variance["Absolute_Variance_vs_Forecast"] / variance["Forecast_Amount"].abs(),
        0,
    )
    variance["Favorable_Variance"] = variance.apply(
        lambda row: row["SAP_Actual_Amount"] - row["Forecast_Amount"]
        if row["Account_Type"] == "Revenue"
        else row["Forecast_Amount"] - row["SAP_Actual_Amount"],
        axis=1,
    )
    variance["Variance_Status"] = np.select(
        [
            variance["Favorable_Variance"] < -50000,
            variance["Favorable_Variance"] > 50000,
            variance["SAP_Actual_Amount"].eq(0),
            variance["Forecast_Amount"].eq(0),
        ],
        ["Unfavorable >50k", "Favorable >50k", "No SAP actual", "No TM1 forecast"],
        default="Within threshold",
    )
    return variance.merge(
        properties[["Property_ID", "Property_Name", "Region", "Asset_Type", "Occupancy_Pct", "Energy_Rating"]],
        on="Property_ID",
        how="left",
    )


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="003F7D")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        max_row = ws.max_row
        max_col = ws.max_column
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        for row in ws.iter_rows(min_row=2, max_row=min(max_row, 2500), max_col=max_col):
            for cell in row:
                cell.border = border
                if isinstance(cell.value, (int, float)):
                    if "Pct" in str(ws.cell(row=1, column=cell.column).value) or "Occupancy" in str(ws.cell(row=1, column=cell.column).value):
                        cell.number_format = "0.0%"
                    else:
                        cell.number_format = "#,##0;(#,##0);-"
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            values = [ws.cell(row=row_idx, column=col_idx).value for row_idx in range(1, min(max_row, 200) + 1)]
            width = min(max(len(str(value)) if value is not None else 0 for value in values) + 2, 42)
            ws.column_dimensions[col_letter].width = max(width, 11)
        if max_row > 1 and max_col > 1:
            table_name = "".join(ch for ch in ws.title if ch.isalnum())[:20] + "Tbl"
            table = Table(displayName=table_name, ref=f"A1:{get_column_letter(max_col)}{max_row}")
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            ws.add_table(table)
    wb.save(path)


def main() -> None:
    accounts = build_accounts()
    cost_centers = build_cost_centers()
    properties = build_properties()
    tm1 = build_tm1(accounts, cost_centers, properties)
    sap = build_sap(accounts, cost_centers, properties, tm1)
    variance = build_variance(accounts, properties, sap, tm1)
    rules = pd.DataFrame(
        [
            ["Revenue", "Higher actual than forecast is favorable. Use Actual - Forecast for favorability.", "Flag unfavorable if revenue shortfall > 50k or > 7%."],
            ["Expense", "Lower actual than forecast is favorable. Use Forecast - Actual for favorability.", "Flag unfavorable if overspend > 50k or > 10%."],
            ["Asset", "Capex must stay inside approved envelope unless approved project acceleration exists.", "Flag overspend > 75k or missing project comment."],
            ["Mixed", "Investigate direction from account policy and journal text.", "Always require owner comment."],
        ],
        columns=["Account_Type", "Variance_Logic", "Threshold_Rule"],
    )
    close_calendar = pd.DataFrame(
        [
            ["Day 0", "2026-05-31", "Subledger close", "AP, AR, RE-FX modules closed"],
            ["Day 1", "2026-06-01", "SAP actual export", "FI/CO actuals exported to close folder"],
            ["Day 2", "2026-06-02", "TM1 forecast freeze", "FP&A freezes Forecast V2 for variance review"],
            ["Day 3", "2026-06-03", "Variance review", "Controllers explain top deviations"],
            ["Day 4", "2026-06-04", "CFO pack", "Executive commentary and decision asks"],
        ],
        columns=["Close_Day", "Date", "Milestone", "Definition_of_Done"],
    )

    sap.to_csv(DATA_DIR / "SAP_Actuals_May2026.csv", index=False, encoding="utf-8-sig")
    tm1.to_csv(DATA_DIR / "TM1_Budget_Forecast_May2026.csv", index=False, encoding="utf-8-sig")
    variance.to_csv(DATA_DIR / "SAP_TM1_Variance_Flat_Table.csv", index=False, encoding="utf-8-sig")
    accounts.to_csv(DATA_DIR / "Account_Mapping_and_Rules.csv", index=False, encoding="utf-8-sig")

    sap_xlsx = DATA_DIR / "SAP_Actuals_May2026.xlsx"
    tm1_xlsx = DATA_DIR / "TM1_Budget_Forecast_May2026.xlsx"
    with pd.ExcelWriter(sap_xlsx, engine="openpyxl") as writer:
        sap.to_excel(writer, sheet_name="SAP_Actuals", index=False)
        accounts[["SAP_GL_Account", "SAP_Account_Name", "Account_Group", "Statement", "Account_Type", "Variance_Direction"]].to_excel(writer, sheet_name="SAP_GL_Accounts", index=False)
        cost_centers[["SAP_Cost_Center", "SAP_Cost_Center_Name", "Department", "Region", "Owner"]].to_excel(writer, sheet_name="SAP_CostCenters", index=False)
        properties.to_excel(writer, sheet_name="Properties", index=False)
        close_calendar.to_excel(writer, sheet_name="Close_Calendar", index=False)
    with pd.ExcelWriter(tm1_xlsx, engine="openpyxl") as writer:
        tm1.to_excel(writer, sheet_name="TM1_Budget_Forecast", index=False)
        accounts[["TM1_Account", "SAP_GL_Account", "SAP_Account_Name", "Account_Group", "Account_Type", "Variance_Direction"]].to_excel(writer, sheet_name="TM1_Account_Mapping", index=False)
        cost_centers[["TM1_Cost_Center", "SAP_Cost_Center", "SAP_Cost_Center_Name", "Department", "Region", "Owner"]].to_excel(writer, sheet_name="TM1_CostCenter_Mapping", index=False)
        rules.to_excel(writer, sheet_name="Variance_Rules", index=False)
        variance.to_excel(writer, sheet_name="Variance_Flat_Table", index=False)

    format_workbook(sap_xlsx)
    format_workbook(tm1_xlsx)

    print(f"SAP rows: {len(sap):,}")
    print(f"TM1 rows: {len(tm1):,}")
    print(f"Variance rows: {len(variance):,}")
    print(f"Output: {DATA_DIR}")
    for file in sorted(DATA_DIR.glob("*")):
        print(f" - {file.name}: {file.stat().st_size / 1024 / 1024:.2f} MB")


if __name__ == "__main__":
    main()
